"""test_excel_parser.py — ExcelParser 单元测试。"""
import openpyxl
import pytest

from backend.rag.preprocessing.ast import walk
from backend.rag.preprocessing.parser.excel_parser import ExcelParser


@pytest.fixture
def sample_xlsx(tmp_path):
    """创建含 2 个 sheet 的 xlsx。"""
    p = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "库存"
    ws1.append(["SKU", "数量", "仓库"])
    ws1.append(["A001", "100", "上海"])
    ws2 = wb.create_sheet("价格")
    ws2.append(["SKU", "价格"])
    ws2.append(["A001", "99.9"])
    wb.save(str(p))
    return str(p)


def test_excel_parser_basic_load(sample_xlsx):
    """每个 sheet 产一个 section（标题=sheet 名）。"""
    ast = ExcelParser().parse(sample_xlsx)
    assert ast.source_file == sample_xlsx
    assert ast.raw_text != ""
    sections = [
        n.text for n in walk(ast.root)
        if n.type == "section" and n.level > 0
    ]
    assert "库存" in sections
    assert "价格" in sections


def test_excel_parser_table_rows_and_text(sample_xlsx):
    """每个 sheet 的表格产 table 节点，rows 有数据，text 非空（避免空 chunk 丢数据）。"""
    ast = ExcelParser().parse(sample_xlsx)
    tables = [n for n in walk(ast.root) if n.type == "table"]
    assert len(tables) == 2
    assert all(t.rows is not None and len(t.rows) >= 2 for t in tables)
    # text 非空，包含单元格内容
    assert any("A001" in t.text for t in tables)


def test_excel_parser_empty_cell_handled(tmp_path):
    """空单元格 → 空字符串，不抛异常。"""
    p = tmp_path / "empty_cell.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", None, "C"])
    ws.append([None, "B", None])
    wb.save(str(p))

    ast = ExcelParser().parse(str(p))
    tables = [n for n in walk(ast.root) if n.type == "table"]
    assert len(tables) == 1
    assert tables[0].rows[0] == ["A", "", "C"]


def test_excel_wired_into_pipeline():
    """.xlsx 接入流水线：parser 注册 + pipeline/loader/indexer 白名单。"""
    from backend.rag.preprocessing.parser import _PARSERS
    from backend.rag.preprocessing.pipeline import _SUPPORTED_EXTS
    from backend.rag.indexing.indexer import IncrementalIndexer

    assert ".xlsx" in _PARSERS
    assert ".xlsx" in _SUPPORTED_EXTS
    assert ".xlsx" in IncrementalIndexer.SUPPORTED_EXTS
