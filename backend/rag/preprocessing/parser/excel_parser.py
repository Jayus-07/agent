"""ExcelParser — openpyxl 解析 Excel → Raw AST。

每个 sheet → section（标题=sheet 名）+ table 节点（rows=二维数组，
text=序列化文本，供 chunk 保留内容，避免空 chunk 丢数据）。
"""
from __future__ import annotations

import openpyxl

from backend.rag.preprocessing.ast import DocumentAST, DocumentNode
from backend.rag.preprocessing.parser.base import BaseDocumentParser
from backend.shared.logger import logger


class ExcelParser(BaseDocumentParser):
    def parse(self, file_path: str) -> DocumentAST:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            logger.error(
                f"[ExcelParser] 打开失败 {file_path}: {type(e).__name__}: {e}"
            )
            return DocumentAST(
                root=DocumentNode(type="section", text="", level=0),
                source_file=file_path,
                raw_text="",
            )

        root = DocumentNode(type="section", text="", level=0)
        raw_lines: list[str] = []

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows: list[list[str]] = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):  # 跳过全空行
                        rows.append(cells)
                if not rows:
                    continue

                section = DocumentNode(type="section", text=sheet_name, level=1)
                # V1.1: 用 NL + CSV 双格式，让 CrossEncoder rerank 能识别表格语义
                from backend.rag.preprocessing.parser._table_nl import make_table_chunk_text
                table_text = make_table_chunk_text(rows, section_title=sheet_name)
                table_node = DocumentNode(type="table", text=table_text, rows=rows)
                section.children.append(table_node)
                root.children.append(section)
                raw_lines.append(f"[{sheet_name}]\n{table_text}")
        finally:
            wb.close()

        raw_text = "\n".join(raw_lines)
        return DocumentAST(root=root, source_file=file_path, raw_text=raw_text)
